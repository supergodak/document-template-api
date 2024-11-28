from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
import openpyxl
import boto3
import re
from io import BytesIO
import logging
from dotenv import load_dotenv
import os

# .envファイルの読み込み
load_dotenv()

# 環境変数の取得
MY_AWS_ACCESS_KEY = os.getenv("MY_AWS_ACCESS_KEY_ID")
MY_AWS_SECRET_KEY = os.getenv("MY_AWS_SECRET_ACCESS_KEY")
MY_AWS_REGION = os.getenv("MY_AWS_REGION", "ap-northeast-1")


app = FastAPI()

# Boto3クライアントの作成
s3_client = boto3.client(
    's3',
    aws_access_key_id=MY_AWS_ACCESS_KEY,
    aws_secret_access_key=MY_AWS_SECRET_KEY,
    region_name=MY_AWS_REGION
)

class InputData(BaseModel):
    bucket_name: str
    object_key: str

# 正規表現でプレースホルダを見つけるパターン
PLACEHOLDER_PATTERN = re.compile(r'\$\{[a-zA-Z0-9_]+\}')

@app.post("/extract_placeholders")
async def extract_placeholders(data: InputData):
    change_key_list = {}

    try:
        # S3からExcelファイルを取得
        response = s3_client.get_object(Bucket=data.bucket_name, Key=data.object_key)
        excel_file = BytesIO(response['Body'].read())
        workbook = openpyxl.load_workbook(excel_file, data_only=False)

        # 各シートからプレースホルダを抽出
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            placeholders = set()

            # セルからプレースホルダを抽出
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        placeholders.update(PLACEHOLDER_PATTERN.findall(cell.value))

            # 図形からプレースホルダを抽出（オプション）
            if sheet._drawing is not None:
                for drawing in sheet._drawing._drawings:
                    if hasattr(drawing, 'text'):
                        placeholders.update(PLACEHOLDER_PATTERN.findall(drawing.text))

            if placeholders:
                change_key_list[sheet_name] = list(placeholders)

        # 成功時のレスポンス
        return {
            "template_id": data.object_key.split('/')[-1].split('.')[0],
            "change_key_list": change_key_list
        }

    except Exception as e:
        logging.error(f"Error processing Excel file: {e}")
        # 失敗時のレスポンス
        return {
            "template_id": data.object_key.split('/')[-1].split('.')[0],
            "error": [str(e)],
            "change_key_list": {}
        }

# ローカル実行用
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)

