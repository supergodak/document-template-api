import requests
import logging
from fastapi import FastAPI, HTTPException, BackgroundTasks
from pydantic import BaseModel
import openpyxl
import boto3
import re
from io import BytesIO
from dotenv import load_dotenv
import os

# ロギングの設定
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# .envファイルの読み込み
load_dotenv()

# 環境変数の取得
AWS_ACCESS_KEY = os.getenv("MY_AWS_ACCESS_KEY_ID")
AWS_SECRET_KEY = os.getenv("MY_AWS_SECRET_ACCESS_KEY")
AWS_REGION = os.getenv("MY_AWS_REGION", "ap-northeast-1")

app = FastAPI()


@app.middleware("http")
async def log_requests(request, call_next):
    logging.info(f"Received request: {request.method} {request.url}")
    response = await call_next(request)
    return response


@app.get("/")
async def root():
    return {"message": "Hello from Vercel!"}


# Boto3クライアントの作成
s3_client = boto3.client('s3',
                         aws_access_key_id=AWS_ACCESS_KEY,
                         aws_secret_access_key=AWS_SECRET_KEY,
                         region_name=AWS_REGION)

# 正規表現でプレースホルダを見つけるパターン
PLACEHOLDER_PATTERN = re.compile(r'\$\{[^}]+\}')


class ExtractPlaceholdersRequest(BaseModel):
    bucket_name: str
    object_key: str


@app.post("/api/extract_placeholders")
async def extract_placeholders(request: ExtractPlaceholdersRequest):
    change_key_list = {}

    try:
        logging.info(
            f"Extracting placeholders from bucket: {request.bucket_name}, object: {request.object_key}"
        )
        # S3からExcelファイルを取得
        response = s3_client.get_object(Bucket=request.bucket_name,
                                        Key=request.object_key)
        excel_file = BytesIO(response['Body'].read())
        workbook = openpyxl.load_workbook(excel_file, data_only=False)

        # 各シートからプレースホルダを抽出
        for sheet_name in workbook.sheetnames:
            logging.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            placeholders = set()

            # セルからプレースホルダを抽出
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        found_placeholders = PLACEHOLDER_PATTERN.findall(
                            cell.value)
                        if found_placeholders:
                            logging.info(
                                f"Found placeholders in cell: {found_placeholders}"
                            )
                        placeholders.update(found_placeholders)

            # 図形からプレースホルダを抽出（オプション）
            if sheet._drawing is not None:
                for drawing in sheet._drawing._drawings:
                    if hasattr(drawing, 'text'):
                        found_placeholders = PLACEHOLDER_PATTERN.findall(
                            drawing.text)
                        if found_placeholders:
                            logging.info(
                                f"Found placeholders in drawing: {found_placeholders}"
                            )
                        placeholders.update(found_placeholders)

            if placeholders:
                change_key_list[sheet_name] = list(placeholders)

        # 成功時のレスポンス
        logging.info(f"Successfully extracted placeholders: {change_key_list}")
        return {
            "template_id": request.object_key.split('/')[-1].split('.')[0],
            "change_key_list": change_key_list
        }

    except Exception as e:
        logging.error(f"Error processing Excel file: {e}")
        # 失敗時のレスポンス
        return {
            "template_id": request.object_key.split('/')[-1].split('.')[0],
            "error": [str(e)],
            "change_key_list": {}
        }


class ReportGenerationRequest(BaseModel):
    output_id: int
    template_id: str
    template_url: str
    template_color: str
    replace_info: dict


# グローバルスコープでバケット名を定義
bucket_name = "my-excel-storage-bucket"


@app.post("/api/generate_report")
async def generate_report(request: ReportGenerationRequest,
                          background_tasks: BackgroundTasks):

    logging.error("ここはgenerate_report")

    # S3 への接続と Excel の取得
    #s3_client = boto3.client('s3')
    s3_client = boto3.client('s3',
                             aws_access_key_id=AWS_ACCESS_KEY,
                             aws_secret_access_key=AWS_SECRET_KEY,
                             region_name=AWS_REGION)

    template_path = request.template_url.split("/")[-1]

    try:
        response = s3_client.get_object(Bucket=bucket_name, Key=template_path)
        excel_file = BytesIO(response['Body'].read())
        workbook = openpyxl.load_workbook(excel_file, data_only=False)

        # 背景タスクとして非同期処理を実行
        background_tasks.add_task(process_excel, workbook,
                                  request.replace_info, request.output_id,
                                  request.template_id)

        # 200 レスポンスをすぐに返す
        response_data = {"message": "Processing started", "status": "success"}
        return response_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


async def process_excel(workbook, replace_info, output_id, template_id):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for key, replacements in replace_info.items():
            for placeholder, values in replacements.items():
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value,
                                      str) and placeholder in cell.value:
                            cell.value = cell.value.replace(
                                placeholder, values[0])  # 最初の置換えのみ適用

    output_file = BytesIO()
    workbook.save(output_file)
    output_file.seek(0)
    save_key = f"processed_template_{output_id}.xlsx"  # セーブするファイル名をここで指定
    s3_client.put_object(Bucket=bucket_name, Key=save_key, Body=output_file)

    report_url = f"https://your-bucket-name.s3.amazonaws.com/{save_key}"  # URLを適切に設定
    await notify_api(output_id, template_id, report_url)


async def notify_api(output_id, template_id, report_url):
    api_url = "https://example.com/notify"  # 通知するAPIのURLを設定
    data = {
        "output_id": output_id,
        "template_id": template_id,
        "report_url": report_url
    }
    response = requests.post(api_url, json=data)
    if response.status_code != 200:
        logging.error(
            f"Failed to notify API: {response.status_code}, {response.text}")


# ローカル実行用
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
